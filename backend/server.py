"""Warehouse Workforce Management API.

FastAPI + MongoDB backend implementing:
- JWT auth with role-based access (manager, asst_manager, document_controller, employee)
- Employee management with admin approval
- Auto-generated rotating shift schedules (2-week cycles)
- Attendance tracking and monthly reports
- Leave management (annual, sick, comp-off, emergency) with coverage validation
"""
from __future__ import annotations

import logging
import os
import uuid
from io import BytesIO
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated, Any, Optional
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, HTTPException, WebSocket, WebSocketDisconnect, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from jose import JWTError, jwt
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from pydantic import BaseModel, EmailStr, Field
from pymongo.errors import DuplicateKeyError

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET_KEY = os.environ["JWT_SECRET_KEY"]
JWT_ALGORITHM = os.environ.get("JWT_ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.environ.get("ACCESS_TOKEN_EXPIRE_MINUTES", "10080"))
APP_TIMEZONE = os.environ.get("APP_TIMEZONE", "Asia/Dubai")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("warehouse")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login", auto_error=False)

ADMIN_ROLES = {"manager", "asst_manager", "document_controller"}
SHIFT_HOURS = {
    "morning": 9,      # 7-16
    "afternoon": 9,    # 12-21
    "night": 9,        # 21-06
    "admin": 9,        # 7:30-16:30
    "sat_day": 12,     # 06-18
    "sat_night": 12,   # 18-06
    "sun_day": 12,
    "sun_night": 12,
    "ega": 9,
    "off": 0,
    "leave": 0,
}
SHIFT_TIMES = {
    "morning": ("07:00", "16:00"),
    "afternoon": ("12:00", "21:00"),
    "night": ("21:00", "06:00"),
    "admin": ("07:30", "16:30"),
    "sat_day": ("06:00", "18:00"),
    "sat_night": ("18:00", "06:00"),
    "sun_day": ("06:00", "18:00"),
    "sun_night": ("18:00", "06:00"),
    "ega": ("07:00", "16:00"),
    "off": ("", ""),
    "leave": ("", ""),
}


def local_now() -> datetime:
    return datetime.now(ZoneInfo(APP_TIMEZONE))


def local_today_str() -> str:
    return local_now().date().isoformat()


class RealtimeHub:
    def __init__(self):
        self.connections: set[WebSocket] = set()

    async def connect(self, websocket: WebSocket) -> None:
        await websocket.accept()
        self.connections.add(websocket)

    def disconnect(self, websocket: WebSocket) -> None:
        self.connections.discard(websocket)

    async def broadcast(self, topic: str, action: str, payload: Optional[dict] = None) -> None:
        stale: list[WebSocket] = []
        message = {
            "topic": topic,
            "action": action,
            "payload": payload or {},
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }
        for websocket in list(self.connections):
            try:
                await websocket.send_json(message)
            except Exception:
                stale.append(websocket)
        for websocket in stale:
            self.disconnect(websocket)


realtime = RealtimeHub()

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    full_name: str
    role: str = "employee"  # employee | manager | asst_manager | document_controller
    team: Optional[str] = None  # "A" | "B" | None (for admins)
    location: str = "warehouse"  # warehouse | ega
    avatar_url: Optional[str] = None


class UserPublic(BaseModel):
    id: str
    email: EmailStr
    full_name: str
    role: str
    status: str  # pending | active | disabled
    team: Optional[str] = None
    location: str = "warehouse"
    default_shift: Optional[str] = None
    avatar_url: Optional[str] = None
    annual_leave_balance: float = 30
    sick_leave_balance: float = 12
    comp_off_balance: float = 0
    created_at: datetime


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserPublic


class UserUpdate(BaseModel):
    email: Optional[EmailStr] = None
    full_name: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    team: Optional[str] = None
    location: Optional[str] = None
    default_shift: Optional[str] = None
    status: Optional[str] = None
    avatar_url: Optional[str] = None
    annual_leave_balance: Optional[float] = None
    sick_leave_balance: Optional[float] = None
    comp_off_balance: Optional[float] = None


class CompOffGrant(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    earned_date: str
    overtime_hours: float = 0
    days: float
    reason: str
    granted_by: str
    granted_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    source_id: Optional[str] = None


class CompOffGrantCreate(BaseModel):
    earned_date: str
    overtime_hours: float = 0
    days: float
    reason: str


class VacationAssign(BaseModel):
    start_date: str
    end_date: str
    reason: str = "Admin assigned vacation"
    leave_type: str = "annual"


class ScheduleEntry(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    avatar_url: Optional[str] = None
    team: Optional[str] = None
    role: Optional[str] = None
    location: Optional[str] = None
    shift_date: str  # YYYY-MM-DD
    shift_type: str  # morning | afternoon | night | admin | sat_day | sat_night | sun_day | sun_night | ega | off | leave
    start_time: str
    end_time: str
    hours: float
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ScheduleCreate(BaseModel):
    user_id: str
    shift_date: str
    shift_type: str
    notes: Optional[str] = None


class GenerateScheduleRequest(BaseModel):
    start_date: str  # Monday YYYY-MM-DD
    weeks: int = 2
    active_saturday_team: str = "A"  # which team works Saturday for week 1
    sunday_team_a_user_id: Optional[str] = None
    sunday_team_b_user_id: Optional[str] = None
    sunday_team_a_shift: str = "sun_day"
    sunday_team_b_shift: str = "sun_night"


class AttendanceMark(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    attendance_date: str  # YYYY-MM-DD
    status: str  # present | absent | late | half_day
    clock_in: Optional[str] = None
    clock_out: Optional[str] = None
    hours_worked: float = 0
    shift_type: Optional[str] = None
    notes: Optional[str] = None
    marked_by: str
    marked_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class AttendanceCreate(BaseModel):
    user_id: Optional[str] = None  # None = self
    attendance_date: str
    status: str = "present"
    clock_in: Optional[str] = None
    clock_out: Optional[str] = None
    hours_worked: Optional[float] = None
    notes: Optional[str] = None


class LeaveRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    leave_type: str  # annual | sick | comp_off | emergency
    start_date: str
    end_date: str
    days: float
    reason: str
    status: str = "pending"  # pending | approved | rejected
    approved_by: Optional[str] = None
    approval_notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LeaveCreate(BaseModel):
    leave_type: str
    start_date: str
    end_date: str
    reason: str


class LeaveAction(BaseModel):
    action: str  # approve | reject
    notes: Optional[str] = None

class SwapRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    requester_id: str
    requester_name: str
    swap_user_id: str
    swap_user_name: str
    shift_date: str
    requester_original_shift: str
    swap_user_original_shift: str
    reason: str
    status: str = "pending_employee_approval"
    employee_decision_at: Optional[datetime] = None
    employee_notes: Optional[str] = None
    admin_decision_at: Optional[datetime] = None
    admin_id: Optional[str] = None
    admin_notes: Optional[str] = None
    executed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class SwapCreate(BaseModel):
    swap_user_id: str
    shift_date: str
    reason: str


class SwapAction(BaseModel):
    action: str  # approve | reject | cancel
    notes: Optional[str] = None

# ---------------------------------------------------------------------------
# Lifespan: db connect + seed
# ---------------------------------------------------------------------------
@asynccontextmanager
async def lifespan(app: FastAPI):
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    app.state.client = client
    app.state.db = db

    await db.users.create_index("email", unique=True)
    await db.schedules.create_index([("user_id", 1), ("shift_date", 1)], unique=True)
    await db.attendance.create_index([("user_id", 1), ("attendance_date", 1)], unique=True)
    await db.comp_off_grants.create_index([("user_id", 1), ("earned_date", -1)])
    await db.comp_off_grants.create_index("source_id", unique=True, sparse=True)
    await db.swap_requests.create_index([("shift_date", 1), ("status", 1)])
    await db.swap_requests.create_index([("requester_id", 1), ("created_at", -1)])
    await db.swap_requests.create_index([("swap_user_id", 1), ("created_at", -1)])
    await db.audit_events.create_index([("entity_type", 1), ("entity_id", 1), ("created_at", -1)])

    await _seed_initial_users(db)
    await _cleanup_placeholder_staff(db)
    await _cleanup_admin_sunday_schedules(db)
    logger.info("Warehouse API ready")
    try:
        yield
    finally:
        client.close()


async def _seed_initial_users(db) -> None:
    """Seed manager/asst/dc only if missing."""
    seeds = [
        # admins
        {
            "email": os.environ["SEED_MANAGER_EMAIL"],
            "password": os.environ["SEED_MANAGER_PASSWORD"],
            "full_name": "Warehouse Manager",
            "role": "manager",
            "team": "A",
            "location": "warehouse",
            "default_shift": "admin",
            "status": "active",
        },
        {
            "email": os.environ["SEED_ASST_EMAIL"],
            "password": os.environ["SEED_ASST_PASSWORD"],
            "full_name": "Assistant Manager",
            "role": "asst_manager",
            "team": "B",
            "location": "warehouse",
            "default_shift": "admin",
            "status": "active",
        },
        {
            "email": os.environ["SEED_DC_EMAIL"],
            "password": os.environ["SEED_DC_PASSWORD"],
            "full_name": "Document Controller",
            "role": "document_controller",
            "team": None,
            "location": "warehouse",
            "default_shift": "admin",
            "status": "active",
        },
    ]
    for s in seeds:
        existing = await db.users.find_one({"email": s["email"]})
        if existing:
            continue
        doc = {
            "_id": str(uuid.uuid4()),
            "email": s["email"],
            "hashed_password": pwd_context.hash(s["password"]),
            "full_name": s["full_name"],
            "role": s["role"],
            "status": s["status"],
            "team": s.get("team"),
            "location": s.get("location", "warehouse"),
            "default_shift": s.get("default_shift"),
            "avatar_url": s.get("avatar_url"),
            "annual_leave_balance": 30,
            "sick_leave_balance": 12,
            "comp_off_balance": 0,
            "created_at": datetime.now(timezone.utc),
        }
        try:
            await db.users.insert_one(doc)
            logger.info("Seeded user: %s (%s)", s["email"], s["role"])
        except DuplicateKeyError:
            pass


async def _cleanup_placeholder_staff(db) -> None:
    """Remove old demo staff that were previously auto-seeded and should not return."""
    placeholder_query = {
        "role": "employee",
        "$or": [
            {"email": {"$in": [
                "staff_a1@warehouse.com",
                "staff_a2@warehouse.com",
                "staff_b1@warehouse.com",
                "staff_b2@warehouse.com",
                "staff_b3@warehouse.com",
                "staff_b4@warehouse.com",
                "testing@warehouse.com",
                "test@warehouse.com",
            ]}},
            {"full_name": {"$regex": r"^(Staff [AB]\d|Testing|Test User)$", "$options": "i"}},
        ],
    }
    docs = await db.users.find(placeholder_query).to_list(100)
    for doc in docs:
        uid = doc["_id"]
        await db.users.delete_one({"_id": uid})
        await db.schedules.delete_many({"user_id": uid})
        await db.attendance.delete_many({"user_id": uid})
        await db.leaves.delete_many({"user_id": uid})
        await db.comp_off_grants.delete_many({"user_id": uid})
    if docs:
        logger.info("Removed %s placeholder/demo staff records", len(docs))


async def _cleanup_admin_sunday_schedules(db) -> None:
    """Existing data cleanup: admin roles must never be scheduled on Sunday."""
    admin_users = await db.users.find({"role": {"$in": list(ADMIN_ROLES)}}).to_list(100)
    cleaned = 0
    for user_doc in admin_users:
        docs = await db.schedules.find({"user_id": user_doc["_id"]}).to_list(1000)
        for sched in docs:
            try:
                shift_day = date.fromisoformat(sched["shift_date"])
            except (KeyError, ValueError):
                continue
            if shift_day.weekday() != 6 or sched.get("shift_type") in ("off", "leave"):
                continue
            await db.schedules.update_one(
                {"user_id": user_doc["_id"], "shift_date": sched["shift_date"]},
                {"$set": {
                    "shift_type": "off",
                    "start_time": "",
                    "end_time": "",
                    "hours": 0,
                    "notes": "Auto-corrected: admin roles are off on Sunday",
                }},
            )
            await _revoke_sunday_comp_off(db, user_doc["_id"], sched["shift_date"])
            cleaned += 1
    if cleaned:
        logger.info("Cleaned %s admin Sunday schedule entries", cleaned)


# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------
app = FastAPI(title="Warehouse Workforce API", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_db():
    return app.state.db


# ---------------------------------------------------------------------------
# Auth utils
# ---------------------------------------------------------------------------
def _create_token(user_id: str, role: str) -> str:
    now = datetime.now(timezone.utc)
    payload = {
        "sub": user_id,
        "role": role,
        "iat": now,
        "exp": now + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)


def _user_public(doc: dict) -> UserPublic:
    return UserPublic(
        id=doc["_id"],
        email=doc["email"],
        full_name=doc["full_name"],
        role=doc["role"],
        status=doc["status"],
        team=doc.get("team"),
        location=doc.get("location", "warehouse"),
        default_shift=doc.get("default_shift"),
        avatar_url=doc.get("avatar_url"),
        annual_leave_balance=doc.get("annual_leave_balance", 30),
        sick_leave_balance=doc.get("sick_leave_balance", 12),
        comp_off_balance=doc.get("comp_off_balance", 0),
        created_at=doc["created_at"],
    )


async def get_current_user(token: Annotated[Optional[str], Depends(oauth2_scheme)], db=Depends(get_db)) -> dict:
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"_id": user_id})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    if user["status"] != "active":
        raise HTTPException(status_code=403, detail="account_not_active")
    return user


def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin role required")
    return user


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------
@app.post("/api/auth/register", response_model=UserPublic)
async def register(payload: UserCreate, db=Depends(get_db)):
    doc = {
        "_id": str(uuid.uuid4()),
        "email": payload.email.lower(),
        "hashed_password": pwd_context.hash(payload.password),
        "full_name": payload.full_name,
        "role": "employee",  # always employee via self-register
        "status": "pending",
        "team": payload.team,
        "location": payload.location,
        "default_shift": None,
        "avatar_url": payload.avatar_url,
        "annual_leave_balance": 30,
        "sick_leave_balance": 12,
        "comp_off_balance": 0,
        "created_at": datetime.now(timezone.utc),
    }
    try:
        await db.users.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(status_code=400, detail="Email already registered")
    await realtime.broadcast("users", "registered", {"user_id": doc["_id"]})
    return _user_public(doc)


@app.post("/api/auth/login", response_model=TokenResponse)
async def login(payload: LoginRequest, db=Depends(get_db)):
    user = await db.users.find_one({"email": payload.email.lower()})
    if not user or not pwd_context.verify(payload.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if user["status"] == "pending":
        raise HTTPException(status_code=403, detail="pending_approval")
    if user["status"] != "active":
        raise HTTPException(status_code=403, detail="account_disabled")
    token = _create_token(user["_id"], user["role"])
    return TokenResponse(access_token=token, user=_user_public(user))


@app.get("/api/auth/me", response_model=UserPublic)
async def me(user: dict = Depends(get_current_user)):
    return _user_public(user)


# ---------------------------------------------------------------------------
# Users (admin)
# ---------------------------------------------------------------------------
@app.get("/api/users", response_model=list[UserPublic])
async def list_users(
    status_filter: Optional[str] = None,
    role: Optional[str] = None,
    location: Optional[str] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    q: dict[str, Any] = {}
    if status_filter:
        q["status"] = status_filter
    if role:
        q["role"] = role
    if location:
        q["location"] = location
    docs = await db.users.find(q).sort("full_name", 1).to_list(500)
    return [_user_public(d) for d in docs]


@app.patch("/api/users/{user_id}", response_model=UserPublic)
async def update_user(user_id: str, payload: UserUpdate, admin: dict = Depends(require_admin), db=Depends(get_db)):
    update = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if "email" in update:
        update["email"] = update["email"].lower()
    password = update.pop("password", None)
    if password is not None:
        if len(password) < 6:
            raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
        update["hashed_password"] = pwd_context.hash(password)
    if not update:
        raise HTTPException(status_code=400, detail="No fields to update")
    try:
        res = await db.users.find_one_and_update(
            {"_id": user_id}, {"$set": update}, return_document=True
        )
    except DuplicateKeyError:
        raise HTTPException(status_code=400, detail="Email already registered")
    if not res:
        raise HTTPException(status_code=404, detail="User not found")
    await realtime.broadcast("users", "updated", {"user_id": user_id})
    return _user_public(res)


@app.post("/api/users/{user_id}/comp-off", response_model=CompOffGrant)
async def grant_comp_off(
    user_id: str,
    payload: CompOffGrantCreate,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    if payload.days <= 0:
        raise HTTPException(status_code=400, detail="Comp-off days must be greater than zero")
    if payload.overtime_hours < 0:
        raise HTTPException(status_code=400, detail="Overtime hours cannot be negative")
    if not payload.reason.strip():
        raise HTTPException(status_code=400, detail="Reason is required")
    try:
        date.fromisoformat(payload.earned_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Earned date must be YYYY-MM-DD")

    target = await db.users.find_one({"_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    grant = CompOffGrant(
        user_id=user_id,
        user_name=target["full_name"],
        earned_date=payload.earned_date,
        overtime_hours=round(payload.overtime_hours, 2),
        days=round(payload.days, 2),
        reason=payload.reason.strip(),
        granted_by=admin["full_name"],
    )
    await db.comp_off_grants.insert_one(grant.dict())
    await db.users.update_one({"_id": user_id}, {"$inc": {"comp_off_balance": grant.days}})
    await realtime.broadcast(
        "users",
        "comp_off_granted",
        {"user_id": user_id, "days": grant.days, "earned_date": grant.earned_date},
    )
    await realtime.broadcast("reports", "comp_off_granted", {"user_id": user_id})
    return grant


@app.get("/api/users/{user_id}/comp-off", response_model=list[CompOffGrant])
async def list_comp_off_grants(
    user_id: str,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    target = await db.users.find_one({"_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    docs = await db.comp_off_grants.find({"user_id": user_id}, {"_id": 0}).sort("earned_date", -1).to_list(100)
    return [CompOffGrant(**d) for d in docs]


@app.post("/api/users/{user_id}/vacation", response_model=LeaveRequest)
async def assign_approved_vacation(
    user_id: str,
    payload: VacationAssign,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    dates = _date_range(payload.start_date, payload.end_date)
    if not dates:
        raise HTTPException(status_code=400, detail="Invalid vacation date range")
    if not payload.reason.strip():
        raise HTTPException(status_code=400, detail="Reason is required")

    target = await db.users.find_one({"_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    days = len(dates)

    lt = payload.leave_type or "annual"
    if lt not in ("annual", "sick", "comp_off", "emergency"):
        raise HTTPException(status_code=400, detail="Invalid leave type")

    # balance check
    if lt == "annual" and target.get("annual_leave_balance", 0) < days:
        raise HTTPException(status_code=400, detail="Insufficient vacation balance")
    elif lt == "sick" and target.get("sick_leave_balance", 0) < days:
        raise HTTPException(status_code=400, detail="Insufficient sick leave balance")
    elif lt == "comp_off" and target.get("comp_off_balance", 0) < days:
        raise HTTPException(status_code=400, detail="Insufficient comp-off balance")

    now = datetime.now(timezone.utc)
    leave = LeaveRequest(
        user_id=user_id,
        user_name=target["full_name"],
        leave_type=lt,
        start_date=payload.start_date,
        end_date=payload.end_date,
        days=days,
        reason=payload.reason.strip(),
        status="approved",
        approved_by=admin["_id"],
        approval_notes="Approved and assigned by admin",
        created_at=now,
        updated_at=now,
    )
    await db.leaves.insert_one(leave.dict())

    # balance reduction
    if lt == "annual":
        await db.users.update_one({"_id": user_id}, {"$inc": {"annual_leave_balance": -days}})
    elif lt == "sick":
        await db.users.update_one({"_id": user_id}, {"$inc": {"sick_leave_balance": -days}})
    elif lt == "comp_off":
        await db.users.update_one({"_id": user_id}, {"$inc": {"comp_off_balance": -days}})
        await _record_comp_off_usage(db, leave.dict(), days, admin["full_name"])
    elif lt == "emergency":
        comp_off_available = max(0, int(target.get("comp_off_balance", 0)))
        comp_off_deduction = min(comp_off_available, days)
        if comp_off_deduction:
            await db.users.update_one({"_id": user_id}, {"$inc": {"comp_off_balance": -comp_off_deduction}})
            await _record_comp_off_usage(db, leave.dict(), comp_off_deduction, admin["full_name"])

    await _apply_approved_leave_to_schedule(db, leave.dict())
    await realtime.broadcast("leaves", "approved", {"leave_id": leave.id, "user_id": user_id})
    await realtime.broadcast(
        "schedules",
        "leave_applied",
        {"leave_id": leave.id, "user_id": user_id, "start_date": leave.start_date, "end_date": leave.end_date},
    )
    await realtime.broadcast("users", "leave_balance_updated", {"leave_id": leave.id, "user_id": user_id})
    await realtime.broadcast("reports", "vacation_assigned", {"leave_id": leave.id, "user_id": user_id})
    return leave


@app.post("/api/users/{user_id}/approve", response_model=UserPublic)
async def approve_user(user_id: str, admin: dict = Depends(require_admin), db=Depends(get_db)):
    res = await db.users.find_one_and_update(
        {"_id": user_id}, {"$set": {"status": "active"}}, return_document=True
    )
    if not res:
        raise HTTPException(status_code=404, detail="User not found")
    await realtime.broadcast("users", "approved", {"user_id": user_id})
    return _user_public(res)


@app.post("/api/users", response_model=UserPublic)
async def create_user(payload: UserCreate, admin: dict = Depends(require_admin), db=Depends(get_db)):
    if payload.role not in {"manager", "asst_manager", "document_controller", "employee"}:
        raise HTTPException(status_code=400, detail="Invalid role")
    doc = {
        "_id": str(uuid.uuid4()),
        "email": payload.email.lower(),
        "hashed_password": pwd_context.hash(payload.password),
        "full_name": payload.full_name,
        "role": payload.role,
        "status": "active",
        "team": payload.team,
        "location": payload.location,
        "default_shift": "admin" if payload.role in ADMIN_ROLES else None,
        "avatar_url": payload.avatar_url,
        "annual_leave_balance": 30,
        "sick_leave_balance": 12,
        "comp_off_balance": 0,
        "created_at": datetime.now(timezone.utc),
    }
    try:
        await db.users.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(status_code=400, detail="Email already registered")
    await realtime.broadcast("users", "created", {"user_id": doc["_id"]})
    return _user_public(doc)


@app.delete("/api/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin), db=Depends(get_db)):
    if user_id == admin["_id"]:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    target = await db.users.find_one({"_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    await db.attendance.delete_many({"user_id": user_id})
    await db.schedules.delete_many({"user_id": user_id})
    await db.leaves.delete_many({"user_id": user_id})
    await db.comp_off_grants.delete_many({"user_id": user_id})
    await db.users.delete_one({"_id": user_id})
    await realtime.broadcast("users", "deleted", {"user_id": user_id})
    await realtime.broadcast("all", "user_data_deleted", {"user_id": user_id})
    return {"deleted": True}


@app.post("/api/admin/reset-operational-data")
async def reset_operational_data(admin: dict = Depends(require_admin), db=Depends(get_db)):
    attendance = await db.attendance.delete_many({})
    schedules = await db.schedules.delete_many({})
    leaves = await db.leaves.delete_many({})
    await realtime.broadcast("all", "operational_data_reset", {})
    return {
        "deleted": {
            "attendance": attendance.deleted_count,
            "schedules": schedules.deleted_count,
            "leaves": leaves.deleted_count,
        }
    }


# ---------------------------------------------------------------------------
# Schedule
# ---------------------------------------------------------------------------
def _hours_for(shift_type: str) -> float:
    return SHIFT_HOURS.get(shift_type, 0)


def _times_for(shift_type: str) -> tuple[str, str]:
    return SHIFT_TIMES.get(shift_type, ("", ""))


def _sunday_comp_source(user_id: str, shift_date: str) -> str:
    return f"sunday-duty:{shift_date}:{user_id}"


async def _grant_sunday_comp_off(db, user_doc: dict, shift_date: str, admin_name: str) -> bool:
    source_id = _sunday_comp_source(user_doc["_id"], shift_date)
    existing = await db.comp_off_grants.find_one({"source_id": source_id})
    if existing:
        return False
    grant = CompOffGrant(
        user_id=user_doc["_id"],
        user_name=user_doc["full_name"],
        earned_date=shift_date,
        overtime_hours=12,
        days=1,
        reason="Sunday duty schedule auto comp off",
        granted_by=admin_name,
        source_id=source_id,
    )
    try:
        await db.comp_off_grants.insert_one(grant.dict())
    except DuplicateKeyError:
        return False
    await db.users.update_one({"_id": user_doc["_id"]}, {"$inc": {"comp_off_balance": 1}})
    return True


async def _revoke_sunday_comp_off(db, user_id: str, shift_date: str) -> bool:
    source_id = _sunday_comp_source(user_id, shift_date)
    result = await db.comp_off_grants.delete_one({"source_id": source_id})
    if result.deleted_count:
        await db.users.update_one({"_id": user_id}, {"$inc": {"comp_off_balance": -1}})
        return True
    return False


async def _record_comp_off_usage(db, leave: dict, days: float, admin_name: str) -> None:
    if days <= 0:
        return
    source_id = f"leave-comp-off:{leave['id']}"
    existing = await db.comp_off_grants.find_one({"source_id": source_id})
    if existing:
        return
    usage = CompOffGrant(
        user_id=leave["user_id"],
        user_name=leave["user_name"],
        earned_date=leave["start_date"],
        overtime_hours=0,
        days=-round(days, 2),
        reason=f"Comp off used for approved {leave['leave_type']} leave ({leave['start_date']} to {leave['end_date']})",
        granted_by=admin_name,
        source_id=source_id,
    )
    try:
        await db.comp_off_grants.insert_one(usage.dict())
    except DuplicateKeyError:
        return


async def _apply_approved_leave_to_schedule(db, leave: dict) -> None:
    applied_dates = _date_range(leave["start_date"], leave["end_date"])
    for d in applied_dates:
        existing_schedule = await db.schedules.find_one({"user_id": leave["user_id"], "shift_date": d})
        if (existing_schedule or {}).get("shift_type") in ("sun_day", "sun_night"):
            await _revoke_sunday_comp_off(db, leave["user_id"], d)
        previous_shift = (existing_schedule or {}).get("shift_type", "unscheduled")
        await db.schedules.update_one(
            {"user_id": leave["user_id"], "shift_date": d},
            {"$set": {
                "id": (existing_schedule or {}).get("id", str(uuid.uuid4())),
                "user_id": leave["user_id"],
                "user_name": leave["user_name"],
                "shift_date": d,
                "shift_type": "leave",
                "start_time": "",
                "end_time": "",
                "hours": 0,
                "notes": f"Leave approved: {leave['leave_type']} ({leave['id']}); previous shift: {previous_shift}",
                "created_at": (existing_schedule or {}).get("created_at", datetime.now(timezone.utc)),
            }},
            upsert=True,
        )


@app.get("/api/schedules", response_model=list[ScheduleEntry])
async def get_schedules(
    start_date: str,
    end_date: str,
    user_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    q = {"shift_date": {"$gte": start_date, "$lte": end_date}}
    if user["role"] not in ADMIN_ROLES and user_id and user_id != user["_id"]:
        raise HTTPException(status_code=403, detail="Cannot view other schedules")
    if user_id:
        q["user_id"] = user_id
    docs = await db.schedules.find(q, {"_id": 0}).sort("shift_date", 1).to_list(2000)
    user_ids = list({d["user_id"] for d in docs})
    user_docs = await db.users.find({"_id": {"$in": user_ids}}, {
        "_id": 1, "avatar_url": 1, "team": 1, "role": 1, "location": 1,
    }).to_list(len(user_ids) or 1)
    user_map = {u["_id"]: u for u in user_docs}
    for d in docs:
        u = user_map.get(d["user_id"], {})
        d["avatar_url"] = u.get("avatar_url")
        d["team"] = u.get("team")
        d["role"] = u.get("role")
        d["location"] = u.get("location")
    return [ScheduleEntry(**d) for d in docs]


@app.post("/api/schedules", response_model=ScheduleEntry)
async def create_or_update_schedule(
    payload: ScheduleCreate, admin: dict = Depends(require_admin), db=Depends(get_db)
):
    u = await db.users.find_one({"_id": payload.user_id})
    if not u:
        raise HTTPException(status_code=404, detail="User not found")
    try:
        shift_day = date.fromisoformat(payload.shift_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="shift_date must be YYYY-MM-DD")
    is_sunday = shift_day.weekday() == 6
    if is_sunday and u["role"] in ADMIN_ROLES and payload.shift_type not in ("off", "leave"):
        raise HTTPException(status_code=400, detail="Manager, Assistant Manager, and Document Controller are off on Sunday")
    if payload.shift_type in ("sun_day", "sun_night"):
        if not is_sunday:
            raise HTTPException(status_code=400, detail="Sunday duty can only be assigned on Sunday")
        if u["role"] in ADMIN_ROLES or u.get("location") == "ega" or u.get("team") not in ("A", "B"):
            raise HTTPException(status_code=400, detail="Sunday duty must be one non-EGA staff member from Team A or Team B")
        team_existing = await db.schedules.find_one({
            "shift_date": payload.shift_date,
            "shift_type": {"$in": ["sun_day", "sun_night"]},
            "user_id": {"$ne": u["_id"]},
        })
        if team_existing:
            existing_user = await db.users.find_one({"_id": team_existing["user_id"]})
            if existing_user and existing_user.get("team") == u.get("team"):
                raise HTTPException(status_code=400, detail=f"Team {u.get('team')} already has Sunday duty assigned")
    existing = await db.schedules.find_one({"user_id": u["_id"], "shift_date": payload.shift_date})
    start, end = _times_for(payload.shift_type)
    entry = ScheduleEntry(
        user_id=u["_id"],
        user_name=u["full_name"],
        shift_date=payload.shift_date,
        shift_type=payload.shift_type,
        start_time=start,
        end_time=end,
        hours=_hours_for(payload.shift_type),
        notes=payload.notes,
    )
    await db.schedules.update_one(
        {"user_id": u["_id"], "shift_date": payload.shift_date},
        {"$set": entry.dict()},
        upsert=True,
    )
    old_sunday = (existing or {}).get("shift_type") in ("sun_day", "sun_night")
    new_sunday = payload.shift_type in ("sun_day", "sun_night")
    comp_changed = False
    if new_sunday:
        comp_changed = await _grant_sunday_comp_off(db, u, payload.shift_date, admin["full_name"])
    elif old_sunday:
        comp_changed = await _revoke_sunday_comp_off(db, u["_id"], payload.shift_date)
    await realtime.broadcast("schedules", "saved", {"user_id": u["_id"], "date": payload.shift_date})
    if comp_changed:
        await realtime.broadcast("users", "comp_off_updated", {"user_id": u["_id"], "date": payload.shift_date})
    return entry


@app.delete("/api/schedules/{user_id}/{shift_date}")
async def delete_schedule(user_id: str, shift_date: str, admin: dict = Depends(require_admin), db=Depends(get_db)):
    existing = await db.schedules.find_one({"user_id": user_id, "shift_date": shift_date})
    await db.schedules.delete_one({"user_id": user_id, "shift_date": shift_date})
    if (existing or {}).get("shift_type") in ("sun_day", "sun_night"):
        if await _revoke_sunday_comp_off(db, user_id, shift_date):
            await realtime.broadcast("users", "comp_off_updated", {"user_id": user_id, "date": shift_date})
    await realtime.broadcast("schedules", "deleted", {"user_id": user_id, "date": shift_date})
    return {"deleted": True}


@app.post("/api/schedules/generate")
async def generate_schedule(
    payload: GenerateScheduleRequest, admin: dict = Depends(require_admin), db=Depends(get_db)
):
    """Auto-generate a 2-week rotating schedule based on each employee's default_shift.

    Rules:
    - Manager/Asst Manager/DC: Mon-Fri admin shift (7:30-16:30)
      - 1st Saturday: Manager off, Asst+DC work admin shift
      - 2nd Saturday: Asst+DC off, Manager works admin shift
    - EGA staff: Mon-Sat ega shift, Sunday off
    - Warehouse staff (default_shift in morning/afternoon/night):
      - Mon-Fri: their default_shift
      - Saturday: if their team == active_saturday_team, work 12-hr shift; else off
      - Sunday: only the selected Team A and Team B users work; everyone else is off
      - Sunday duty automatically earns 1 comp-off day for the selected staff
    """
    try:
        start = datetime.strptime(payload.start_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="start_date must be YYYY-MM-DD")
    if start.weekday() != 0:
        raise HTTPException(status_code=400, detail="start_date must be a Monday")

    days = payload.weeks * 7
    users = await db.users.find({"status": "active"}).to_list(200)
    end_date = (start + timedelta(days=days - 1)).isoformat()
    approved_leaves = await db.leaves.find({
        "status": "approved",
        "start_date": {"$lte": end_date},
        "end_date": {"$gte": payload.start_date},
    }, {"_id": 0}).to_list(5000)
    approved_leave_days: set[tuple[str, str]] = set()
    for lv in approved_leaves:
        for leave_day in _date_range(lv["start_date"], lv["end_date"]):
            if payload.start_date <= leave_day <= end_date:
                approved_leave_days.add((lv["user_id"], leave_day))

    saturday_teams = [payload.active_saturday_team]
    saturday_teams.append("B" if payload.active_saturday_team == "A" else "A")

    users_by_id = {u["_id"]: u for u in users}
    sunday_a = users_by_id.get(payload.sunday_team_a_user_id or "")
    sunday_b = users_by_id.get(payload.sunday_team_b_user_id or "")
    if payload.sunday_team_a_user_id and (not sunday_a or sunday_a.get("team") != "A" or sunday_a.get("role") in ADMIN_ROLES or sunday_a.get("location") == "ega"):
        raise HTTPException(status_code=400, detail="Sunday Team A staff must be an active non-admin Team A user")
    if payload.sunday_team_b_user_id and (not sunday_b or sunday_b.get("team") != "B" or sunday_b.get("role") in ADMIN_ROLES or sunday_b.get("location") == "ega"):
        raise HTTPException(status_code=400, detail="Sunday Team B staff must be an active non-admin Team B user")
    if payload.sunday_team_a_shift not in ("sun_day", "sun_night"):
        raise HTTPException(status_code=400, detail="Sunday Team A shift must be sun_day or sun_night")
    if payload.sunday_team_b_shift not in ("sun_day", "sun_night"):
        raise HTTPException(status_code=400, detail="Sunday Team B shift must be sun_day or sun_night")
    sunday_assignments = {
        payload.sunday_team_a_user_id: payload.sunday_team_a_shift,
        payload.sunday_team_b_user_id: payload.sunday_team_b_shift,
    }
    sunday_assignments.pop(None, None)
    sunday_assignments.pop("", None)

    generated = 0
    leave_preserved = 0
    comp_off_added = 0
    comp_off_removed = 0
    for i in range(days):
        day = start + timedelta(days=i)
        weekday = day.weekday()  # 0=Mon ... 6=Sun
        week_idx = i // 7
        date_str = day.isoformat()
        sat_team = saturday_teams[week_idx % 2]

        for u in users:
            role = u["role"]
            team = u.get("team")
            loc = u.get("location", "warehouse")
            shift_default = u.get("default_shift")
            shift_type = "off"

            if (u["_id"], date_str) in approved_leave_days:
                shift_type = "leave"
                leave_preserved += 1
            elif role in ADMIN_ROLES:
                if weekday <= 4:  # Mon-Fri
                    shift_type = "admin"
                elif weekday == 5:  # Saturday
                    if role == "manager":
                        shift_type = "off" if week_idx % 2 == 0 else "admin"
                    else:  # asst_manager / DC
                        shift_type = "admin" if week_idx % 2 == 0 else "off"
                else:
                    shift_type = "off"
            elif loc == "ega":
                shift_type = "ega" if weekday <= 5 else "off"
            else:
                if weekday <= 4:
                    shift_type = shift_default or "morning"
                elif weekday == 5:
                    if team == sat_team:
                        shift_type = "sat_day"
                    else:
                        shift_type = "off"
                else:  # Sunday
                    shift_type = sunday_assignments.get(u["_id"], "off")

            start_t, end_t = _times_for(shift_type)
            entry = {
                "id": str(uuid.uuid4()),
                "user_id": u["_id"],
                "user_name": u["full_name"],
                "shift_date": date_str,
                "shift_type": shift_type,
                "start_time": start_t,
                "end_time": end_t,
                "hours": _hours_for(shift_type),
                "notes": None,
                "created_at": datetime.now(timezone.utc),
            }
            existing = await db.schedules.find_one({"user_id": u["_id"], "shift_date": date_str})
            await db.schedules.update_one(
                {"user_id": u["_id"], "shift_date": date_str},
                {"$set": entry},
                upsert=True,
            )
            old_sunday = (existing or {}).get("shift_type") in ("sun_day", "sun_night")
            new_sunday = shift_type in ("sun_day", "sun_night")
            if new_sunday:
                if await _grant_sunday_comp_off(db, u, date_str, admin["full_name"]):
                    comp_off_added += 1
            elif old_sunday:
                if await _revoke_sunday_comp_off(db, u["_id"], date_str):
                    comp_off_removed += 1
            generated += 1

    await realtime.broadcast("schedules", "generated", {"start_date": payload.start_date, "days": days})
    if comp_off_added or comp_off_removed:
        await realtime.broadcast("users", "comp_off_updated", {"added": comp_off_added, "removed": comp_off_removed})
    return {
        "generated": generated,
        "days": days,
        "leave_preserved": leave_preserved,
        "comp_off_added": comp_off_added,
        "comp_off_removed": comp_off_removed,
    }


# ---------------------------------------------------------------------------
# Attendance
# ---------------------------------------------------------------------------
def _compute_hours(clock_in: Optional[str], clock_out: Optional[str]) -> float:
    if not clock_in or not clock_out:
        return 0
    try:
        ci = datetime.strptime(clock_in, "%H:%M")
        co = datetime.strptime(clock_out, "%H:%M")
        if co < ci:
            co += timedelta(days=1)
        return round((co - ci).total_seconds() / 3600, 2)
    except ValueError:
        return 0


async def _matching_leave_for_date(db, user_id: str, attendance_date: str, leave_type: Optional[str] = None) -> Optional[dict]:
    query: dict[str, Any] = {
        "user_id": user_id,
        "start_date": {"$lte": attendance_date},
        "end_date": {"$gte": attendance_date},
        "status": {"$in": ["pending", "approved"]},
    }
    if leave_type:
        query["leave_type"] = leave_type
    return await db.leaves.find_one(query)


async def _approve_manual_sick_day(db, target: dict, attendance_date: str, admin: dict, notes: Optional[str]) -> dict:
    existing = await _matching_leave_for_date(db, target["_id"], attendance_date, "sick")
    if existing and existing.get("status") == "approved":
        return existing
    if target.get("sick_leave_balance", 0) < 1:
        raise HTTPException(status_code=400, detail="Insufficient sick leave balance")

    now = datetime.now(timezone.utc)
    if existing:
        await db.leaves.update_one(
            {"id": existing["id"]},
            {"$set": {
                "status": "approved",
                "approved_by": admin["_id"],
                "approval_notes": notes or "Approved from manual attendance entry",
                "updated_at": now,
            }},
        )
        leave = await db.leaves.find_one({"id": existing["id"]})
    else:
        leave_model = LeaveRequest(
            user_id=target["_id"],
            user_name=target["full_name"],
            leave_type="sick",
            start_date=attendance_date,
            end_date=attendance_date,
            days=1,
            reason=notes or "Sick day recorded from manual attendance",
            status="approved",
            approved_by=admin["_id"],
            approval_notes="Approved from manual attendance entry",
        )
        leave = leave_model.dict()
        await db.leaves.insert_one(leave)

    await db.users.update_one({"_id": target["_id"]}, {"$inc": {"sick_leave_balance": -1}})
    await _apply_approved_leave_to_schedule(db, leave)
    await _audit_event(db, "leave", leave["id"], "manual_sick_approved", admin, {"date": attendance_date})
    await realtime.broadcast("leaves", "approved", {"leave_id": leave["id"], "user_id": target["_id"]})
    await realtime.broadcast("users", "leave_balance_updated", {"leave_id": leave["id"], "user_id": target["_id"]})
    await realtime.broadcast("schedules", "leave_applied", {"leave_id": leave["id"], "user_id": target["_id"], "start_date": attendance_date, "end_date": attendance_date})
    return leave


async def _reconcile_previous_day_leave(db, user: dict) -> list[dict]:
    previous_date = (local_now().date() - timedelta(days=1)).isoformat()
    schedule = await db.schedules.find_one({
        "user_id": user["_id"],
        "shift_date": previous_date,
        "shift_type": {"$nin": ["off", "leave"]},
    })
    attendance = await db.attendance.find_one({"user_id": user["_id"], "attendance_date": previous_date})
    formal_leave = await _matching_leave_for_date(db, user["_id"], previous_date)

    if schedule and not attendance and not formal_leave:
        record = AttendanceMark(
            user_id=user["_id"],
            user_name=user["full_name"],
            attendance_date=previous_date,
            status="leave",
            hours_worked=0,
            shift_type=schedule.get("shift_type"),
            notes="Auto-marked because attendance was not recorded. Formal leave request required.",
            marked_by="system",
        )
        await db.attendance.update_one(
            {"user_id": user["_id"], "attendance_date": previous_date},
            {"$set": record.dict()},
            upsert=True,
        )
        attendance = record.dict()
        await realtime.broadcast("attendance", "auto_leave", {"user_id": user["_id"], "date": previous_date})

    if not attendance or attendance.get("status") not in ("leave", "sick", "comp_off"):
        return []

    expected_type = {"sick": "sick", "comp_off": "comp_off"}.get(attendance.get("status"))
    matching_leave = await _matching_leave_for_date(db, user["_id"], previous_date, expected_type)
    if matching_leave:
        return []
    return [{
        "attendance_date": previous_date,
        "attendance_status": attendance.get("status", "leave"),
        "suggested_leave_type": expected_type or "annual",
        "message": f"Attendance for {previous_date} is marked as {attendance.get('status', 'leave').replace('_', ' ')}. Submit the formal leave request through the app.",
    }]


@app.post("/api/attendance", response_model=AttendanceMark)
async def mark_attendance(
    payload: AttendanceCreate, user: dict = Depends(get_current_user), db=Depends(get_db)
):
    target_id = payload.user_id or user["_id"]
    if target_id != user["_id"] and user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Cannot mark for others")
    target = await db.users.find_one({"_id": target_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    sched = await db.schedules.find_one({"user_id": target_id, "shift_date": payload.attendance_date})
    existing = await db.attendance.find_one({"user_id": target_id, "attendance_date": payload.attendance_date}) or {}

    if payload.status not in ("present", "late", "absent", "half_day", "leave", "sick", "comp_off"):
        raise HTTPException(status_code=400, detail="Invalid attendance status")
    if payload.status == "sick" and user["role"] in ADMIN_ROLES:
        await _approve_manual_sick_day(db, target, payload.attendance_date, user, payload.notes)

    # Partial merge so clock-in and clock-out can be submitted independently
    clock_in = payload.clock_in if payload.clock_in is not None else existing.get("clock_in")
    clock_out = payload.clock_out if payload.clock_out is not None else existing.get("clock_out")

    hours = payload.hours_worked
    if hours is None:
        if clock_in and clock_out:
            hours = _compute_hours(clock_in, clock_out)
        elif sched and payload.status == "present" and not clock_in:
            hours = sched.get("hours", 0)
        else:
            hours = existing.get("hours_worked", 0)

    record = AttendanceMark(
        user_id=target_id,
        user_name=target["full_name"],
        attendance_date=payload.attendance_date,
        status=payload.status,
        clock_in=clock_in,
        clock_out=clock_out,
        hours_worked=hours,
        shift_type=(sched or {}).get("shift_type"),
        notes=payload.notes if payload.notes is not None else existing.get("notes"),
        marked_by=user["_id"],
    )
    await db.attendance.update_one(
        {"user_id": target_id, "attendance_date": payload.attendance_date},
        {"$set": record.dict()},
        upsert=True,
    )
    await realtime.broadcast("attendance", "saved", {"user_id": target_id, "date": payload.attendance_date})
    return record


@app.post("/api/attendance/clock-in", response_model=AttendanceMark)
async def clock_in_now(user: dict = Depends(get_current_user), db=Depends(get_db)):
    """Record current time as clock-in for today."""
    now_str = local_now().strftime("%H:%M")
    payload = AttendanceCreate(
        attendance_date=local_today_str(),
        status="present",
        clock_in=now_str,
    )
    return await mark_attendance(payload, user, db)


@app.post("/api/attendance/clock-out", response_model=AttendanceMark)
async def clock_out_now(user: dict = Depends(get_current_user), db=Depends(get_db)):
    """Record current time as clock-out for today (hours auto-computed)."""
    now_str = local_now().strftime("%H:%M")
    payload = AttendanceCreate(
        attendance_date=local_today_str(),
        status="present",
        clock_out=now_str,
    )
    return await mark_attendance(payload, user, db)


@app.get("/api/attendance", response_model=list[AttendanceMark])
async def get_attendance(
    start_date: str,
    end_date: str,
    user_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    q = {"attendance_date": {"$gte": start_date, "$lte": end_date}}
    if user["role"] not in ADMIN_ROLES:
        q["user_id"] = user["_id"]
    elif user_id:
        q["user_id"] = user_id
    docs = await db.attendance.find(q, {"_id": 0}).sort("attendance_date", 1).to_list(2000)
    return [AttendanceMark(**d) for d in docs]


@app.delete("/api/attendance/{user_id}/{attendance_date}")
async def delete_attendance(
    user_id: str,
    attendance_date: str,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    result = await db.attendance.delete_one({"user_id": user_id, "attendance_date": attendance_date})
    if result.deleted_count:
        await realtime.broadcast("attendance", "deleted", {"user_id": user_id, "date": attendance_date})
    return {"deleted": result.deleted_count}


@app.get("/api/attendance/monthly/{user_id}/{year}/{month}")
async def monthly_attendance(
    user_id: str, year: int, month: int,
    user: dict = Depends(get_current_user), db=Depends(get_db),
):
    if user["role"] not in ADMIN_ROLES and user_id != user["_id"]:
        raise HTTPException(status_code=403, detail="Cannot view others")
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    docs = await db.attendance.find(
        {"user_id": user_id, "attendance_date": {"$gte": start.isoformat(), "$lte": end.isoformat()}},
        {"_id": 0},
    ).to_list(500)
    total = sum(d.get("hours_worked", 0) for d in docs)
    present = sum(1 for d in docs if d["status"] == "present")
    absent = sum(1 for d in docs if d["status"] == "absent")
    late = sum(1 for d in docs if d["status"] == "late")
    half = sum(1 for d in docs if d["status"] == "half_day")
    return {
        "user_id": user_id,
        "year": year,
        "month": month,
        "total_hours": round(total, 2),
        "present": present,
        "absent": absent,
        "late": late,
        "half_day": half,
        "records": docs,
    }



# ---------------------------------------------------------------------------
# Swap requests and audit trail
# ---------------------------------------------------------------------------
async def _audit_event(db, entity_type: str, entity_id: str, action: str, actor: dict, details: Optional[dict] = None) -> None:
    await db.audit_events.insert_one({
        "id": str(uuid.uuid4()),
        "entity_type": entity_type,
        "entity_id": entity_id,
        "action": action,
        "actor_id": actor.get("_id"),
        "actor_name": actor.get("full_name", "System"),
        "details": details or {},
        "created_at": datetime.now(timezone.utc),
    })


def _swap_public(doc: dict) -> SwapRequest:
    clean = {k: v for k, v in doc.items() if k != "_id"}
    return SwapRequest(**clean)


async def _available_swap_candidates(db, requester_id: str, shift_date: str) -> list[dict]:
    requester_schedule = await db.schedules.find_one({"user_id": requester_id, "shift_date": shift_date})
    if not requester_schedule:
        return []
    on_leave = await db.leaves.find({
        "status": "approved",
        "start_date": {"$lte": shift_date},
        "end_date": {"$gte": shift_date},
    }).to_list(500)
    unavailable = {lv["user_id"] for lv in on_leave}
    schedules = await db.schedules.find({
        "shift_date": shift_date,
        "user_id": {"$ne": requester_id},
        "shift_type": {"$nin": ["off", "leave"]},
    }, {"_id": 0}).to_list(500)
    candidates = []
    for schedule in schedules:
        if schedule["user_id"] in unavailable:
            continue
        if schedule.get("shift_type") == requester_schedule.get("shift_type"):
            continue
        candidates.append({
            "user_id": schedule["user_id"],
            "user_name": schedule.get("user_name", ""),
            "avatar_url": schedule.get("avatar_url"),
            "team": schedule.get("team"),
            "current_shift": schedule.get("shift_type"),
            "requested_shift": requester_schedule.get("shift_type"),
        })
    return candidates


async def _execute_swap(db, swap: dict, admin: dict) -> None:
    requester_schedule = await db.schedules.find_one({
        "user_id": swap["requester_id"], "shift_date": swap["shift_date"]
    })
    target_schedule = await db.schedules.find_one({
        "user_id": swap["swap_user_id"], "shift_date": swap["shift_date"]
    })
    if not requester_schedule or not target_schedule:
        raise HTTPException(status_code=409, detail="One of the scheduled duties no longer exists")
    if requester_schedule.get("shift_type") != swap.get("requester_original_shift") or target_schedule.get("shift_type") != swap.get("swap_user_original_shift"):
        raise HTTPException(status_code=409, detail="A duty changed after this swap was requested. Cancel it and create a new request.")

    requester_shift = target_schedule["shift_type"]
    target_shift = requester_schedule["shift_type"]
    requester_times = SHIFT_TIMES[requester_shift]
    target_times = SHIFT_TIMES[target_shift]
    now = datetime.now(timezone.utc)

    await db.schedules.update_one(
        {"user_id": swap["requester_id"], "shift_date": swap["shift_date"]},
        {"$set": {
            "shift_type": requester_shift,
            "start_time": requester_times[0],
            "end_time": requester_times[1],
            "hours": SHIFT_HOURS[requester_shift],
            "notes": f"Executed swap {swap['id']} with {swap['swap_user_name']}",
        }},
    )
    await db.schedules.update_one(
        {"user_id": swap["swap_user_id"], "shift_date": swap["shift_date"]},
        {"$set": {
            "shift_type": target_shift,
            "start_time": target_times[0],
            "end_time": target_times[1],
            "hours": SHIFT_HOURS[target_shift],
            "notes": f"Executed swap {swap['id']} with {swap['requester_name']}",
        }},
    )
    await db.attendance.update_one(
        {"user_id": swap["requester_id"], "attendance_date": swap["shift_date"]},
        {"$set": {"shift_type": requester_shift, "notes": f"Schedule updated by swap {swap['id']}"}},
    )
    await db.attendance.update_one(
        {"user_id": swap["swap_user_id"], "attendance_date": swap["shift_date"]},
        {"$set": {"shift_type": target_shift, "notes": f"Schedule updated by swap {swap['id']}"}},
    )
    await db.swap_requests.update_one(
        {"id": swap["id"]},
        {"$set": {
            "status": "executed",
            "admin_id": admin["_id"],
            "admin_decision_at": now,
            "executed_at": now,
            "updated_at": now,
        }},
    )
    await _audit_event(db, "swap", swap["id"], "executed", admin, {
        "shift_date": swap["shift_date"],
        "requester_new_shift": requester_shift,
        "swap_user_new_shift": target_shift,
    })
    await realtime.broadcast("schedules", "swap_executed", {"swap_id": swap["id"], "date": swap["shift_date"]})
    await realtime.broadcast("reports", "swap_executed", {"swap_id": swap["id"]})


@app.get("/api/swaps/candidates")
async def swap_candidates(
    shift_date: str,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    return await _available_swap_candidates(db, user["_id"], shift_date)


@app.post("/api/swaps", response_model=SwapRequest)
async def create_swap_request(
    payload: SwapCreate,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    if payload.swap_user_id == user["_id"]:
        raise HTTPException(status_code=400, detail="Select another employee")
    requester_schedule = await db.schedules.find_one({"user_id": user["_id"], "shift_date": payload.shift_date})
    target_schedule = await db.schedules.find_one({"user_id": payload.swap_user_id, "shift_date": payload.shift_date})
    target_user = await db.users.find_one({"_id": payload.swap_user_id, "status": "active"})
    if not requester_schedule:
        raise HTTPException(status_code=400, detail="You have no scheduled duty on this date")
    if not target_schedule or not target_user:
        raise HTTPException(status_code=400, detail="Selected employee has no available duty on this date")
    if requester_schedule["shift_type"] in ("off", "leave") or target_schedule["shift_type"] in ("off", "leave"):
        raise HTTPException(status_code=400, detail="Off or leave duties cannot be swapped")
    if requester_schedule["shift_type"] == target_schedule["shift_type"]:
        raise HTTPException(status_code=400, detail="Select an employee from another shift")
    coverage = await _check_coverage(db, [payload.shift_date], user["_id"])
    if coverage.get(payload.shift_date, {}).get("ok", True):
        raise HTTPException(
            status_code=400,
            detail="Swap is available only when an otherwise valid leave request would reduce minimum shift coverage",
        )
    conflict = await db.swap_requests.find_one({
        "shift_date": payload.shift_date,
        "$or": [
            {"requester_id": {"$in": [user["_id"], payload.swap_user_id]}},
            {"swap_user_id": {"$in": [user["_id"], payload.swap_user_id]}},
        ],
        "status": {"$in": ["pending_employee_approval", "pending_admin_approval"]},
    })
    if conflict:
        raise HTTPException(status_code=409, detail="A pending swap already involves one of these employees on this date")

    swap = SwapRequest(
        requester_id=user["_id"],
        requester_name=user["full_name"],
        swap_user_id=target_user["_id"],
        swap_user_name=target_user["full_name"],
        shift_date=payload.shift_date,
        requester_original_shift=requester_schedule["shift_type"],
        swap_user_original_shift=target_schedule["shift_type"],
        reason=payload.reason,
    )
    await db.swap_requests.insert_one(swap.dict())
    await _audit_event(db, "swap", swap.id, "requested", user, swap.dict())
    await realtime.broadcast("swaps", "requested", {"swap_id": swap.id, "swap_user_id": payload.swap_user_id})
    return swap


@app.get("/api/swaps", response_model=list[SwapRequest])
async def list_swaps(
    status_filter: Optional[str] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    query: dict[str, Any] = {}
    if status_filter:
        query["status"] = status_filter
    if user["role"] not in ADMIN_ROLES:
        query["$or"] = [{"requester_id": user["_id"]}, {"swap_user_id": user["_id"]}]
    docs = await db.swap_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [_swap_public(doc) for doc in docs]


@app.post("/api/swaps/{swap_id}/employee-action", response_model=SwapRequest)
async def employee_swap_action(
    swap_id: str,
    payload: SwapAction,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    swap = await db.swap_requests.find_one({"id": swap_id})
    if not swap:
        raise HTTPException(status_code=404, detail="Swap request not found")
    if swap["swap_user_id"] != user["_id"]:
        raise HTTPException(status_code=403, detail="Only the selected employee can respond")
    if swap["status"] != "pending_employee_approval":
        raise HTTPException(status_code=400, detail="Swap is not awaiting employee approval")
    if payload.action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="action must be approve or reject")
    now = datetime.now(timezone.utc)
    status_value = "pending_admin_approval" if payload.action == "approve" else "rejected"
    await db.swap_requests.update_one({"id": swap_id}, {"$set": {
        "status": status_value,
        "employee_decision_at": now,
        "employee_notes": payload.notes,
        "updated_at": now,
    }})
    await _audit_event(db, "swap", swap_id, f"employee_{payload.action}", user, {"notes": payload.notes})
    await realtime.broadcast("swaps", status_value, {"swap_id": swap_id})
    return _swap_public(await db.swap_requests.find_one({"id": swap_id}, {"_id": 0}))


@app.post("/api/swaps/{swap_id}/admin-action", response_model=SwapRequest)
async def admin_swap_action(
    swap_id: str,
    payload: SwapAction,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    swap = await db.swap_requests.find_one({"id": swap_id})
    if not swap:
        raise HTTPException(status_code=404, detail="Swap request not found")
    if swap["status"] != "pending_admin_approval":
        raise HTTPException(status_code=400, detail="Swap is not awaiting admin approval")
    if payload.action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="action must be approve or reject")
    if payload.action == "approve":
        await _execute_swap(db, swap, admin)
    else:
        now = datetime.now(timezone.utc)
        await db.swap_requests.update_one({"id": swap_id}, {"$set": {
            "status": "rejected",
            "admin_id": admin["_id"],
            "admin_decision_at": now,
            "admin_notes": payload.notes,
            "updated_at": now,
        }})
        await _audit_event(db, "swap", swap_id, "admin_rejected", admin, {"notes": payload.notes})
        await realtime.broadcast("swaps", "rejected", {"swap_id": swap_id})
    return _swap_public(await db.swap_requests.find_one({"id": swap_id}, {"_id": 0}))


@app.post("/api/swaps/{swap_id}/cancel", response_model=SwapRequest)
async def cancel_swap(
    swap_id: str,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    swap = await db.swap_requests.find_one({"id": swap_id})
    if not swap or swap["requester_id"] != user["_id"]:
        raise HTTPException(status_code=404, detail="Swap request not found")
    if swap["status"] not in ("pending_employee_approval", "pending_admin_approval"):
        raise HTTPException(status_code=400, detail="Swap can no longer be cancelled")
    now = datetime.now(timezone.utc)
    await db.swap_requests.update_one({"id": swap_id}, {"$set": {"status": "cancelled", "updated_at": now}})
    await _audit_event(db, "swap", swap_id, "cancelled", user)
    await realtime.broadcast("swaps", "cancelled", {"swap_id": swap_id})
    return _swap_public(await db.swap_requests.find_one({"id": swap_id}, {"_id": 0}))

# ---------------------------------------------------------------------------
# Leaves
# ---------------------------------------------------------------------------
def _date_range(start: str, end: str) -> list[str]:
    try:
        s = datetime.strptime(start, "%Y-%m-%d").date()
        e = datetime.strptime(end, "%Y-%m-%d").date()
    except ValueError:
        return []
    if e < s:
        return []
    return [(s + timedelta(days=i)).isoformat() for i in range((e - s).days + 1)]


async def _check_coverage(db, dates: list[str], excluding_user: str) -> dict:
    """Return per-date warehouse coverage status (excluding admins/EGA).

    If no schedule has been generated for a given date yet, that date is treated
    as 'ok' (cannot block leave on dates with no roster).
    """
    coverage = {}
    for d in dates:
        schedules = await db.schedules.find({"shift_date": d}).to_list(200)
        if not schedules:
            coverage[d] = {"morning": 0, "afternoon": 0, "night": 0, "ok": True, "no_schedule": True}
            continue
        leaves = await db.leaves.find({
            "status": "approved",
            "start_date": {"$lte": d},
            "end_date": {"$gte": d},
        }).to_list(100)
        on_leave_ids = {lv["user_id"] for lv in leaves}

        counts = {"morning": 0, "afternoon": 0, "night": 0}
        affected_schedule = next((s for s in schedules if s["user_id"] == excluding_user), None)
        affected_shift = (affected_schedule or {}).get("shift_type")
        for s in schedules:
            if s["user_id"] == excluding_user:
                continue
            if s["user_id"] in on_leave_ids:
                continue
            if s["shift_type"] in counts:
                counts[s["shift_type"]] += 1
        minimum = COVERAGE_MIN.get(affected_shift, 0)
        ok = affected_shift not in counts or counts[affected_shift] >= minimum
        coverage[d] = {**counts, "affected_shift": affected_shift, "required": minimum, "ok": ok}
    return coverage


@app.post("/api/leaves", response_model=LeaveRequest)
async def apply_leave(
    payload: LeaveCreate, user: dict = Depends(get_current_user), db=Depends(get_db)
):
    dates = _date_range(payload.start_date, payload.end_date)
    if not dates:
        raise HTTPException(status_code=400, detail="Invalid date range")
    days = len(dates)

    # balance check
    if payload.leave_type == "annual" and user.get("annual_leave_balance", 0) < days:
        raise HTTPException(status_code=400, detail="Insufficient annual leave balance")
    if payload.leave_type == "sick" and user.get("sick_leave_balance", 0) < days:
        raise HTTPException(status_code=400, detail="Insufficient sick leave balance")
    if payload.leave_type == "comp_off" and user.get("comp_off_balance", 0) < days:
        raise HTTPException(status_code=400, detail="Insufficient comp-off balance")

    # coverage check (skip for emergency)
    if payload.leave_type != "emergency" and user.get("location") != "ega" and user["role"] not in ADMIN_ROLES:
        coverage = await _check_coverage(db, dates, user["_id"])
        bad = [d for d, c in coverage.items() if not c["ok"]]
        if bad:
            candidate_map = {}
            for bad_date in bad:
                candidate_map[bad_date] = await _available_swap_candidates(db, user["_id"], bad_date)
            raise HTTPException(
                status_code=409,
                detail={
                    "error": "coverage_insufficient",
                    "message": "Normal leave would reduce the shift below minimum manpower.",
                    "dates": bad,
                    "coverage": coverage,
                    "options": ["emergency_leave", "swap_request"],
                    "swap_candidates": candidate_map,
                },
            )

    leave = LeaveRequest(
        user_id=user["_id"],
        user_name=user["full_name"],
        leave_type=payload.leave_type,
        start_date=payload.start_date,
        end_date=payload.end_date,
        days=days,
        reason=payload.reason,
    )
    await db.leaves.insert_one(leave.dict())
    await realtime.broadcast("leaves", "requested", {"leave_id": leave.id, "user_id": user["_id"]})
    return leave


@app.get("/api/leaves", response_model=list[LeaveRequest])
async def list_leaves(
    status_filter: Optional[str] = None,
    user_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    q: dict[str, Any] = {}
    if status_filter:
        q["status"] = status_filter
    if user["role"] not in ADMIN_ROLES:
        q["user_id"] = user["_id"]
    elif user_id:
        q["user_id"] = user_id
    docs = await db.leaves.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return [LeaveRequest(**d) for d in docs]


@app.post("/api/leaves/{leave_id}/action", response_model=LeaveRequest)
async def act_on_leave(
    leave_id: str, payload: LeaveAction, admin: dict = Depends(require_admin), db=Depends(get_db)
):
    if payload.action not in ("approve", "reject"):
        raise HTTPException(status_code=400, detail="action must be approve or reject")
    leave = await db.leaves.find_one({"id": leave_id})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")
    if leave["status"] != "pending":
        raise HTTPException(status_code=400, detail="Already processed")

    new_status = "approved" if payload.action == "approve" else "rejected"
    await db.leaves.update_one(
        {"id": leave_id},
        {"$set": {
            "status": new_status,
            "approved_by": admin["_id"],
            "approval_notes": payload.notes,
            "updated_at": datetime.now(timezone.utc),
        }},
    )
    # On approve: decrement balance + mark schedule as leave
    if new_status == "approved":
        field_map = {
            "annual": "annual_leave_balance",
            "sick": "sick_leave_balance",
            "comp_off": "comp_off_balance",
        }
        f = field_map.get(leave["leave_type"])
        if f:
            await db.users.update_one({"_id": leave["user_id"]}, {"$inc": {f: -leave["days"]}})
            if leave["leave_type"] == "comp_off":
                await _record_comp_off_usage(db, leave, leave["days"], admin["full_name"])
        elif leave["leave_type"] == "emergency":
            leave_user = await db.users.find_one({"_id": leave["user_id"]}, {"comp_off_balance": 1})
            comp_off_available = max(0, int((leave_user or {}).get("comp_off_balance", 0)))
            comp_off_deduction = min(comp_off_available, leave["days"])
            if comp_off_deduction:
                await db.users.update_one(
                    {"_id": leave["user_id"]},
                    {"$inc": {"comp_off_balance": -comp_off_deduction}},
                )
                await _record_comp_off_usage(db, leave, comp_off_deduction, admin["full_name"])
        await _apply_approved_leave_to_schedule(db, leave)
        await _audit_event(db, "leave", leave_id, "approved", admin, {"leave_type": leave["leave_type"], "days": leave["days"], "notes": payload.notes})

    if new_status == "rejected":
        await _audit_event(db, "leave", leave_id, "rejected", admin, {"notes": payload.notes})
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    await realtime.broadcast("leaves", new_status, {"leave_id": leave_id, "user_id": leave["user_id"]})
    if new_status == "approved":
        await realtime.broadcast(
            "schedules",
            "leave_applied",
            {
                "leave_id": leave_id,
                "user_id": leave["user_id"],
                "start_date": leave["start_date"],
                "end_date": leave["end_date"],
            },
        )
        await realtime.broadcast("users", "leave_balance_updated", {"leave_id": leave_id, "user_id": leave["user_id"]})
    return LeaveRequest(**leave)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@app.get("/api/dashboard")
async def dashboard(user: dict = Depends(get_current_user), db=Depends(get_db)):
    leave_request_reminders = await _reconcile_previous_day_leave(db, user)
    today = local_today_str()
    today_schedule = await db.schedules.find_one({"user_id": user["_id"], "shift_date": today}, {"_id": 0})
    today_attendance_docs = await db.attendance.find(
        {
            "attendance_date": today,
            "status": {"$in": ["present", "late", "half_day"]},
        },
        {"_id": 0, "user_id": 1},
    ).to_list(1000)
    today_signed_in = len({a["user_id"] for a in today_attendance_docs})
    today_leave_docs = await db.leaves.find(
        {
            "status": "approved",
            "start_date": {"$lte": today},
            "end_date": {"$gte": today},
        },
        {"_id": 0, "user_id": 1, "leave_type": 1},
    ).to_list(1000)
    today_sick = len({lv["user_id"] for lv in today_leave_docs if lv.get("leave_type") == "sick"})
    today_comp_off = len({lv["user_id"] for lv in today_leave_docs if lv.get("leave_type") == "comp_off"})
    today_on_leave = len({lv["user_id"] for lv in today_leave_docs})
    today_schedules = await db.schedules.find({
        "shift_date": today,
        "shift_type": {"$nin": ["off", "leave"]},
    }, {"_id": 0}).sort([("start_time", 1), ("user_name", 1)]).to_list(1000)
    today_attendance_by_user = {
        item["user_id"]: item
        for item in await db.attendance.find({"attendance_date": today}, {"_id": 0}).to_list(1000)
    }
    duty_user_ids = list({schedule["user_id"] for schedule in today_schedules})
    duty_users = {
        item["_id"]: item
        for item in await db.users.find(
            {"_id": {"$in": duty_user_ids}},
            {"_id": 1, "full_name": 1, "avatar_url": 1, "team": 1, "role": 1},
        ).to_list(1000)
    } if duty_user_ids else {}
    today_duty_holders = []
    for schedule in today_schedules:
        duty_user = duty_users.get(schedule["user_id"], {})
        attendance_item = today_attendance_by_user.get(schedule["user_id"])
        is_present = bool(attendance_item and attendance_item.get("status") in ("present", "late", "half_day"))
        today_duty_holders.append({
            "user_id": schedule["user_id"],
            "user_name": duty_user.get("full_name") or schedule.get("user_name", ""),
            "avatar_url": duty_user.get("avatar_url") or schedule.get("avatar_url"),
            "team": duty_user.get("team") or schedule.get("team"),
            "role": duty_user.get("role") or schedule.get("role"),
            "shift_type": schedule.get("shift_type"),
            "start_time": schedule.get("start_time", ""),
            "end_time": schedule.get("end_time", ""),
            "attendance_status": attendance_item.get("status") if attendance_item else None,
            "clock_in": attendance_item.get("clock_in") if attendance_item else None,
            "is_present": is_present,
        })
    # this month attendance summary
    start_month = date.today().replace(day=1).isoformat()
    att_docs = await db.attendance.find(
        {"user_id": user["_id"], "attendance_date": {"$gte": start_month, "$lte": today}},
        {"_id": 0},
    ).to_list(50)
    hours_this_month = sum(a.get("hours_worked", 0) for a in att_docs)
    present_days = sum(1 for a in att_docs if a["status"] == "present")
    today_user_attendance = next((a for a in att_docs if a.get("attendance_date") == today), None)
    has_punched_in_today = bool(today_user_attendance)
    pending_leaves = await db.leaves.count_documents({"user_id": user["_id"], "status": "pending"})

    summary = {
        "today_schedule": today_schedule,
        "today_signed_in": today_signed_in,
        "today_sick": today_sick,
        "today_comp_off": today_comp_off,
        "today_on_leave": today_on_leave,
        "today_duty_holders": today_duty_holders,
        "leave_request_reminders": leave_request_reminders,
        "hours_this_month": round(hours_this_month, 2),
        "present_days_this_month": present_days,
        "pending_leaves": pending_leaves,
        "annual_leave_balance": user.get("annual_leave_balance", 0),
        "sick_leave_balance": user.get("sick_leave_balance", 0),
        "comp_off_balance": user.get("comp_off_balance", 0),
        "has_punched_in_today": has_punched_in_today,
    }

    if user["role"] in ADMIN_ROLES:
        pending_users = await db.users.count_documents({"status": "pending"})
        pending_all_leaves = await db.leaves.count_documents({"status": "pending"})
        total_active = await db.users.count_documents({"status": "active"})
        # today coverage
        today_sched = await db.schedules.find({"shift_date": today}).to_list(50)
        counts = {"morning": 0, "afternoon": 0, "night": 0, "admin": 0, "ega": 0}
        for s in today_sched:
            if s["shift_type"] in counts:
                counts[s["shift_type"]] += 1
        summary["admin"] = {
            "pending_user_approvals": pending_users,
            "pending_leave_approvals": pending_all_leaves,
            "total_active_users": total_active,
            "today_coverage": counts,
            "today_signed_in": today_signed_in,
            "today_sick": today_sick,
            "today_comp_off": today_comp_off,
            "today_on_leave": today_on_leave,
            "today_duty_holders": today_duty_holders,
        }
    return summary


@app.get("/api/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}


@app.websocket("/api/realtime/ws")
async def realtime_ws(websocket: WebSocket, token: str, db=Depends(get_db)):
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        user = await db.users.find_one({"_id": user_id}) if user_id else None
        if not user or user.get("status") != "active":
            await websocket.close(code=1008)
            return
    except JWTError:
        await websocket.close(code=1008)
        return

    await realtime.connect(websocket)
    await websocket.send_json({"topic": "all", "action": "connected"})
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        realtime.disconnect(websocket)


# ---------------------------------------------------------------------------
# Calendar / Command Center
# ---------------------------------------------------------------------------
COVERAGE_MIN = {"morning": 3, "afternoon": 2, "night": 2}


@app.get("/api/calendar/month")
async def calendar_month(
    year: int,
    month: int,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    """Aggregated monthly command-center view for ALL users.

    Returns per-day:
      - list of schedule entries grouped by shift_type
      - approved + pending leaves that touch this day
      - coverage counts vs minimum (morning≥3, afternoon≥2, night≥2)
      - status: ok | warn | critical (critical = below_min)

    Plus a monthly comparison summary.
    """
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="month must be 1-12")
    start = date(year, month, 1)
    end = date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1) - timedelta(days=1)
    start_str, end_str = start.isoformat(), end.isoformat()

    schedules = await db.schedules.find(
        {"shift_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}
    ).sort("shift_date", 1).to_list(5000)
    leaves = await db.leaves.find({
        "$and": [
            {"start_date": {"$lte": end_str}},
            {"end_date": {"$gte": start_str}},
            {"status": {"$in": ["approved", "pending"]}},
        ]
    }, {"_id": 0}).to_list(500)
    swaps = await db.swap_requests.find({
        "shift_date": {"$gte": start_str, "$lte": end_str},
        "status": {"$ne": "cancelled"},
    }, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find(
        {"attendance_date": {"$gte": start_str, "$lte": end_str}}, {"_id": 0}
    ).sort([("attendance_date", 1), ("user_name", 1)]).to_list(10000)

    days: dict[str, dict] = {}
    cursor = start
    while cursor <= end:
        d = cursor.isoformat()
        days[d] = {
            "date": d,
            "weekday": cursor.weekday(),  # 0=Mon
            "shifts": {k: [] for k in ["morning", "afternoon", "night", "admin", "ega",
                                       "sat_day", "sat_night", "sun_day", "sun_night", "off", "leave"]},
            "leaves": [],
            "swaps": [],
            "attendance": [],
            "roster": [],
            "roster_summary": {"scheduled": 0, "finished": 0, "clocked_in": 0, "marked": 0, "missing": 0},
            "attendance_summary": {"present": 0, "late": 0, "absent": 0, "half_day": 0, "total": 0},
            "coverage": {"morning": 0, "afternoon": 0, "night": 0},
            "coverage_if_pending_approved": {"morning": 0, "afternoon": 0, "night": 0},
            "pending_leave_impact": {"morning": 0, "afternoon": 0, "night": 0},
            "status": "ok",
            "pending_status": "ok",
        }
        cursor += timedelta(days=1)

    # Bucket schedules by date+shift
    for s in schedules:
        d = s["shift_date"]
        if d not in days:
            continue
        st = s.get("shift_type", "off")
        days[d]["shifts"].setdefault(st, []).append({
            "user_id": s["user_id"],
            "user_name": s["user_name"],
            "start_time": s.get("start_time", ""),
            "end_time": s.get("end_time", ""),
        })
        if st in days[d]["coverage"]:
            days[d]["coverage"][st] += 1

    # Bucket leaves into the dates they overlap (dedup by user_id+leave_id)
    seen_per_day: dict[str, set] = {}
    for lv in leaves:
        ls = max(start, datetime.strptime(lv["start_date"], "%Y-%m-%d").date())
        le = min(end, datetime.strptime(lv["end_date"], "%Y-%m-%d").date())
        cur = ls
        while cur <= le:
            d = cur.isoformat()
            if d in days:
                key = f"{lv['user_id']}:{lv['leave_type']}:{lv['status']}"
                seen = seen_per_day.setdefault(d, set())
                if key not in seen:
                    seen.add(key)
                    days[d]["leaves"].append({
                        "user_id": lv["user_id"],
                        "user_name": lv["user_name"],
                        "leave_type": lv["leave_type"],
                        "status": lv["status"],
                    })
                    # Approved leaves reduce coverage (only once per user per day)
                    if lv["status"] == "approved":
                        for sk in days[d]["coverage"].keys():
                            for entry in days[d]["shifts"].get(sk, []):
                                if entry["user_id"] == lv["user_id"]:
                                    days[d]["coverage"][sk] = max(0, days[d]["coverage"][sk] - 1)
                                    break
            cur += timedelta(days=1)

    # Project coverage if pending leave requests are approved.
    for info in days.values():
        info["coverage_if_pending_approved"] = dict(info["coverage"])
    for lv in leaves:
        if lv["status"] != "pending":
            continue
        ls = max(start, datetime.strptime(lv["start_date"], "%Y-%m-%d").date())
        le = min(end, datetime.strptime(lv["end_date"], "%Y-%m-%d").date())
        cur = ls
        while cur <= le:
            d = cur.isoformat()
            if d in days:
                for sk in days[d]["coverage_if_pending_approved"].keys():
                    for entry in days[d]["shifts"].get(sk, []):
                        if entry["user_id"] == lv["user_id"]:
                            days[d]["coverage_if_pending_approved"][sk] = max(0, days[d]["coverage_if_pending_approved"][sk] - 1)
                            days[d]["pending_leave_impact"][sk] += 1
                            break
            cur += timedelta(days=1)

    # Bucket swaps by duty date.
    for swap in swaps:
        d = swap.get("shift_date")
        if d not in days:
            continue
        days[d]["swaps"].append({
            "id": swap["id"],
            "requester_id": swap["requester_id"],
            "requester_name": swap["requester_name"],
            "swap_user_id": swap["swap_user_id"],
            "swap_user_name": swap["swap_user_name"],
            "requester_original_shift": swap["requester_original_shift"],
            "swap_user_original_shift": swap["swap_user_original_shift"],
            "status": swap["status"],
            "reason": swap.get("reason", ""),
        })

    # Bucket attendance by marked date so the command center reflects real logs.
    attendance_statuses = {"present", "late", "absent", "half_day"}
    for rec in attendance:
        d = rec["attendance_date"]
        if d not in days:
            continue
        status_value = rec.get("status", "")
        days[d]["attendance"].append({
            "user_id": rec["user_id"],
            "user_name": rec.get("user_name", ""),
            "status": status_value,
            "clock_in": rec.get("clock_in"),
            "clock_out": rec.get("clock_out"),
            "hours_worked": rec.get("hours_worked", 0),
            "shift_type": rec.get("shift_type"),
        })
        days[d]["attendance_summary"]["total"] += 1
        if status_value in attendance_statuses:
            days[d]["attendance_summary"][status_value] += 1

    attendance_by_day_user = {
        (rec.get("attendance_date"), rec.get("user_id")): rec
        for rec in attendance
    }
    scheduled_by_day_user: set[tuple[str, str]] = set()
    for s in schedules:
        d = s["shift_date"]
        uid = s["user_id"]
        if d not in days:
            continue
        scheduled_by_day_user.add((d, uid))
        rec = attendance_by_day_user.get((d, uid))
        has_in = bool(rec and rec.get("clock_in"))
        has_out = bool(rec and rec.get("clock_out"))
        if rec and rec.get("status") == "absent":
            log_state = "absent"
        elif has_in and has_out:
            log_state = "finished"
        elif has_in:
            log_state = "clocked_in"
        elif rec:
            log_state = "marked"
        else:
            log_state = "missing"
        days[d]["roster"].append({
            "user_id": uid,
            "user_name": s.get("user_name", ""),
            "shift_type": s.get("shift_type"),
            "start_time": s.get("start_time", ""),
            "end_time": s.get("end_time", ""),
            "scheduled_hours": s.get("hours", 0),
            "attendance_status": rec.get("status") if rec else None,
            "clock_in": rec.get("clock_in") if rec else None,
            "clock_out": rec.get("clock_out") if rec else None,
            "hours_worked": rec.get("hours_worked", 0) if rec else 0,
            "log_state": log_state,
        })
        days[d]["roster_summary"]["scheduled"] += 1
        if log_state in days[d]["roster_summary"]:
            days[d]["roster_summary"][log_state] += 1

    # Include unscheduled attendance too, so manual/admin logs are not invisible.
    for rec in attendance:
        d = rec["attendance_date"]
        uid = rec["user_id"]
        if d not in days or (d, uid) in scheduled_by_day_user:
            continue
        has_in = bool(rec.get("clock_in"))
        has_out = bool(rec.get("clock_out"))
        log_state = "finished" if has_in and has_out else "clocked_in" if has_in else "marked"
        days[d]["roster"].append({
            "user_id": uid,
            "user_name": rec.get("user_name", ""),
            "shift_type": rec.get("shift_type") or "unscheduled",
            "start_time": "",
            "end_time": "",
            "scheduled_hours": 0,
            "attendance_status": rec.get("status"),
            "clock_in": rec.get("clock_in"),
            "clock_out": rec.get("clock_out"),
            "hours_worked": rec.get("hours_worked", 0),
            "log_state": log_state,
        })
        if log_state in days[d]["roster_summary"]:
            days[d]["roster_summary"][log_state] += 1

    for info in days.values():
        info["roster"].sort(key=lambda row: (
            row.get("start_time") or "99:99",
            row.get("user_name") or "",
        ))

    # Compute status per day (only count weekdays mon-fri for strict critical)
    monthly_critical = 0
    monthly_warn = 0
    total_scheduled_hours = 0.0
    for s in schedules:
        total_scheduled_hours += s.get("hours", 0)

    for d, info in days.items():
        cov = info["coverage"]
        is_weekday = info["weekday"] <= 4
        if is_weekday:
            below = (
                cov["morning"] < COVERAGE_MIN["morning"]
                or cov["afternoon"] < COVERAGE_MIN["afternoon"]
                or cov["night"] < COVERAGE_MIN["night"]
            )
            if below:
                info["status"] = "critical"
                monthly_critical += 1
            elif (cov["morning"] == COVERAGE_MIN["morning"]
                  or cov["afternoon"] == COVERAGE_MIN["afternoon"]
                  or cov["night"] == COVERAGE_MIN["night"]):
                info["status"] = "warn"
                monthly_warn += 1
            pending_cov = info["coverage_if_pending_approved"]
            pending_below = (
                pending_cov["morning"] < COVERAGE_MIN["morning"]
                or pending_cov["afternoon"] < COVERAGE_MIN["afternoon"]
                or pending_cov["night"] < COVERAGE_MIN["night"]
            )
            if pending_below:
                info["pending_status"] = "critical"
            elif (pending_cov["morning"] == COVERAGE_MIN["morning"]
                  or pending_cov["afternoon"] == COVERAGE_MIN["afternoon"]
                  or pending_cov["night"] == COVERAGE_MIN["night"]):
                info["pending_status"] = "warn"
        else:
            # weekends: only flag if there's literally no coverage
            total_weekend = sum(cov.values())
            if total_weekend == 0 and not info["leaves"]:
                info["status"] = "warn"

    # Comparison summary
    active_users_count = await db.users.count_documents({"status": "active"})
    approved_leaves_total = sum(1 for lv in leaves if lv["status"] == "approved")
    pending_leaves_total = sum(1 for lv in leaves if lv["status"] == "pending")
    pending_swaps_total = sum(1 for swap in swaps if swap["status"] in ("pending_employee_approval", "pending_admin_approval"))
    executed_swaps_total = sum(1 for swap in swaps if swap["status"] == "executed")
    marked_attendance_total = len(attendance)

    return {
        "range": {"start_date": start_str, "end_date": end_str, "year": year, "month": month},
        "minimum_coverage": COVERAGE_MIN,
        "days": list(days.values()),
        "summary": {
            "total_active_staff": active_users_count,
            "total_scheduled_entries": len(schedules),
            "total_scheduled_hours": round(total_scheduled_hours, 2),
            "approved_leaves": approved_leaves_total,
            "pending_leaves": pending_leaves_total,
            "pending_swaps": pending_swaps_total,
            "executed_swaps": executed_swaps_total,
            "marked_attendance": marked_attendance_total,
            "critical_days": monthly_critical,
            "warn_days": monthly_warn,
        },
    }


@app.get("/api/admin/source-zip")
async def download_source_zip(token: Optional[str] = None, admin_user: Optional[dict] = None, db=Depends(get_db)):
    """Download the complete project source as a zip file.

    Accepts either:
    - A standard Bearer token in Authorization header (via Depends), OR
    - A `token` query param (so the link can be opened directly in a browser).
    Admin role required.
    """
    user = None
    if token:
        try:
            payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
            uid = payload.get("sub")
            user = await db.users.find_one({"_id": uid}) if uid else None
        except JWTError:
            user = None
    if not user or user.get("role") not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Admin token required (?token=...)")
    zip_path = ROOT_DIR / "downloads" / "warehouse-app.zip"
    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="Source zip not built yet")
    return FileResponse(
        path=str(zip_path),
        filename="warehouse-app.zip",
        media_type="application/zip",
    )


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------
@app.get("/api/reports/employee/{user_id}")
async def employee_report(
    user_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(get_current_user),
    db=Depends(get_db),
):
    """Aggregated per-employee report. Defaults to current month if no range given."""
    if user["role"] not in ADMIN_ROLES and user_id != user["_id"]:
        raise HTTPException(status_code=403, detail="Cannot view others")
    target = await db.users.find_one({"_id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    today = date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    # Attendance
    att_docs = await db.attendance.find(
        {"user_id": user_id, "attendance_date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0},
    ).sort("attendance_date", 1).to_list(500)
    total_hours = round(sum(a.get("hours_worked", 0) for a in att_docs), 2)
    present_days = sum(1 for a in att_docs if a["status"] == "present")
    late_days = sum(1 for a in att_docs if a["status"] == "late")
    absent_days = sum(1 for a in att_docs if a["status"] == "absent")
    half_days = sum(1 for a in att_docs if a["status"] == "half_day")

    # Scheduled vs marked
    sched_docs = await db.schedules.find(
        {"user_id": user_id, "shift_date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0},
    ).to_list(500)
    scheduled_work_days = sum(1 for s in sched_docs if s["shift_type"] not in ("off", "leave"))
    scheduled_hours = round(sum(s.get("hours", 0) for s in sched_docs), 2)

    # Leaves
    leave_docs = await db.leaves.find(
        {"user_id": user_id, "start_date": {"$lte": end_date}, "end_date": {"$gte": start_date}},
        {"_id": 0},
    ).to_list(200)
    leave_summary = {
        "annual": {"taken": 0, "pending": 0},
        "sick": {"taken": 0, "pending": 0},
        "comp_off": {"taken": 0, "pending": 0},
        "emergency": {"taken": 0, "pending": 0},
    }
    for lv in leave_docs:
        t = lv["leave_type"]
        if t not in leave_summary:
            continue
        if lv["status"] == "approved":
            leave_summary[t]["taken"] += lv["days"]
        elif lv["status"] == "pending":
            leave_summary[t]["pending"] += lv["days"]

    return {
        "user": _user_public(target).dict(),
        "range": {"start_date": start_date, "end_date": end_date},
        "attendance": {
            "total_hours": total_hours,
            "present_days": present_days,
            "late_days": late_days,
            "absent_days": absent_days,
            "half_days": half_days,
            "scheduled_work_days": scheduled_work_days,
            "scheduled_hours": scheduled_hours,
            "records": att_docs,
        },
        "leaves": {
            "summary": leave_summary,
            "balances": {
                "annual": target.get("annual_leave_balance", 0),
                "sick": target.get("sick_leave_balance", 0),
                "comp_off": target.get("comp_off_balance", 0),
            },
            "records": leave_docs,
        },
    }


@app.get("/api/reports/attendance/all")
async def all_attendance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """Admin report for all marked attendance in a date range."""
    today = date.today()
    if not start_date:
        start_date = today.replace(day=1).isoformat()
    if not end_date:
        end_date = today.isoformat()

    docs = await db.attendance.find(
        {"attendance_date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0},
    ).sort([("attendance_date", 1), ("user_name", 1)]).to_list(10000)

    by_user: dict[str, dict[str, Any]] = {}
    totals = {
        "records": len(docs),
        "total_hours": 0.0,
        "present_days": 0,
        "late_days": 0,
        "absent_days": 0,
        "half_days": 0,
    }
    status_fields = {
        "present": "present_days",
        "late": "late_days",
        "absent": "absent_days",
        "half_day": "half_days",
    }
    for rec in docs:
        hours = float(rec.get("hours_worked", 0) or 0)
        status_value = rec.get("status", "")
        totals["total_hours"] += hours
        if status_value in status_fields:
            totals[status_fields[status_value]] += 1

        user_id = rec["user_id"]
        summary = by_user.setdefault(user_id, {
            "user_id": user_id,
            "user_name": rec.get("user_name", ""),
            "records": 0,
            "total_hours": 0.0,
            "present_days": 0,
            "late_days": 0,
            "absent_days": 0,
            "half_days": 0,
        })
        summary["records"] += 1
        summary["total_hours"] += hours
        if status_value in status_fields:
            summary[status_fields[status_value]] += 1

    totals["total_hours"] = round(totals["total_hours"], 2)
    users = []
    for summary in by_user.values():
        summary["total_hours"] = round(summary["total_hours"], 2)
        users.append(summary)
    users.sort(key=lambda item: item["user_name"])

    return {
        "range": {"start_date": start_date, "end_date": end_date},
        "totals": totals,
        "users": users,
        "records": docs,
    }


@app.get("/api/reports/export")
async def export_report_data(
    start_date: str,
    end_date: str,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """Admin data bundle for frontend Excel-compatible export."""
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    if end < start:
        raise HTTPException(status_code=400, detail="End date must be after start date")
    if (end - start).days > 370:
        raise HTTPException(status_code=400, detail="Export range cannot exceed 370 days")

    users = await db.users.find({"status": "active"}).sort("full_name", 1).to_list(500)
    attendance = await db.attendance.find(
        {"attendance_date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0},
    ).sort([("attendance_date", 1), ("user_name", 1)]).to_list(20000)
    schedules = await db.schedules.find(
        {"shift_date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0},
    ).sort([("shift_date", 1), ("user_name", 1)]).to_list(20000)
    leaves = await db.leaves.find(
        {"start_date": {"$lte": end_date}, "end_date": {"$gte": start_date}},
        {"_id": 0},
    ).sort([("start_date", 1), ("user_name", 1)]).to_list(5000)
    swaps = await db.swap_requests.find(
        {"shift_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort([("shift_date", 1), ("requester_name", 1)]).to_list(5000)
    comp_off_grants = await db.comp_off_grants.find(
        {"earned_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort([("earned_date", 1), ("user_name", 1)]).to_list(5000)

    user_rows = []
    for user_doc in users:
        uid = user_doc["_id"]
        user_att = [a for a in attendance if a["user_id"] == uid]
        user_leaves = [lv for lv in leaves if lv["user_id"] == uid]
        user_rows.append({
            "user_id": uid,
            "user_name": user_doc["full_name"],
            "email": user_doc["email"],
            "role": user_doc["role"],
            "team": user_doc.get("team"),
            "location": user_doc.get("location", "warehouse"),
            "attendance_records": len(user_att),
            "present": sum(1 for a in user_att if a["status"] == "present"),
            "late": sum(1 for a in user_att if a["status"] == "late"),
            "absent": sum(1 for a in user_att if a["status"] == "absent"),
            "half_day": sum(1 for a in user_att if a["status"] == "half_day"),
            "total_hours": round(sum(a.get("hours_worked", 0) for a in user_att), 2),
            "annual_leave": sum(lv.get("days", 0) for lv in user_leaves if lv["leave_type"] == "annual" and lv["status"] == "approved"),
            "sick_leave": sum(lv.get("days", 0) for lv in user_leaves if lv["leave_type"] == "sick" and lv["status"] == "approved"),
            "vacation_leave": sum(lv.get("days", 0) for lv in user_leaves if lv["leave_type"] == "annual" and lv["status"] == "approved"),
            "comp_off_leave": sum(lv.get("days", 0) for lv in user_leaves if lv["leave_type"] == "comp_off" and lv["status"] == "approved"),
            "emergency_leave": sum(lv.get("days", 0) for lv in user_leaves if lv["leave_type"] == "emergency" and lv["status"] == "approved"),
            "pending_leave": sum(lv.get("days", 0) for lv in user_leaves if lv["status"] == "pending"),
            "annual_balance": user_doc.get("annual_leave_balance", 0),
            "sick_balance": user_doc.get("sick_leave_balance", 0),
            "comp_off_balance": user_doc.get("comp_off_balance", 0),
            "swap_duty_days": sum(1 for swap in swaps if swap.get("status") == "executed" and uid in (swap.get("requester_id"), swap.get("swap_user_id"))),
        })

    return {
        "range": {"start_date": start_date, "end_date": end_date},
        "users": user_rows,
        "attendance": attendance,
        "schedules": schedules,
        "leaves": leaves,
        "swaps": swaps,
        "comp_off_grants": comp_off_grants,
    }

@app.get("/api/reports/export.xlsx")
async def export_management_workbook(
    start_date: str,
    end_date: str,
    admin: dict = Depends(require_admin),
    db=Depends(get_db),
):
    """Generate a management-ready Excel workbook from live operational data."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Dates must be YYYY-MM-DD")
    if end < start or (end - start).days > 370:
        raise HTTPException(status_code=400, detail="Export range must be between 1 and 370 days")

    users = await db.users.find({"status": "active"}).sort("full_name", 1).to_list(1000)
    attendance = await db.attendance.find({"attendance_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(30000)
    schedules = await db.schedules.find({"shift_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(30000)
    leaves = await db.leaves.find({"start_date": {"$lte": end_date}, "end_date": {"$gte": start_date}}, {"_id": 0}).to_list(10000)
    swaps = await db.swap_requests.find({"shift_date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}).to_list(10000)
    comp_off = await db.comp_off_grants.find({}, {"_id": 0}).to_list(10000)

    wb = Workbook()
    wb.remove(wb.active)
    colors_x = {
        "header": "111827", "header_text": "FFFFFF", "duty": "F8FAFC",
        "leave": "FECACA", "vacation": "BFDBFE", "comp": "BBF7D0",
        "swap": "E9D5FF", "pending": "FDE68A", "danger": "EF4444", "green": "10B981",
    }
    thin = Side(style="thin", color="D1D5DB")

    def setup_sheet(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=colors_x["header"])
            cell.font = Font(color=colors_x["header_text"], bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.sheet_view.showGridLines = False

    def autofit(ws, maximum=32):
        for column in ws.columns:
            letter = get_column_letter(column[0].column)
            width = min(maximum, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            ws.column_dimensions[letter].width = width

    user_map = {u["_id"]: u for u in users}
    att_by_key = {(a["user_id"], a["attendance_date"]): a for a in attendance}
    sched_by_key = {(s["user_id"], s["shift_date"]): s for s in schedules}
    approved_leaves = [lv for lv in leaves if lv.get("status") == "approved"]
    date_list = [(start + timedelta(days=i)).isoformat() for i in range((end - start).days + 1)]

    summary_rows = []
    for u in users:
        uid = u["_id"]
        user_att = [a for a in attendance if a["user_id"] == uid]
        user_sched = [s for s in schedules if s["user_id"] == uid]
        user_leaves = [lv for lv in approved_leaves if lv["user_id"] == uid]
        executed_swaps = [s for s in swaps if s.get("status") == "executed" and uid in (s.get("requester_id"), s.get("swap_user_id"))]
        summary_rows.append([
            u.get("full_name", ""), uid, u.get("team") or "", u.get("role", ""), u.get("default_shift") or "",
            sum(1 for s in user_sched if s.get("shift_type") not in ("off", "leave")),
            sum(lv.get("days", 0) for lv in user_leaves),
            sum(lv.get("days", 0) for lv in user_leaves if lv.get("leave_type") == "annual"),
            sum(lv.get("days", 0) for lv in user_leaves if lv.get("leave_type") == "comp_off"),
            len(executed_swaps),
            sum(1 for a in user_att if a.get("status") == "absent"),
            round(sum(float(a.get("hours_worked", 0) or 0) for a in user_att), 2),
        ])

    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False
    dash["A1"] = "WAREHOUSE WORKFORCE DASHBOARD"
    dash["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    dash["A1"].fill = PatternFill("solid", fgColor=colors_x["header"])
    dash.merge_cells("A1:F2")
    dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
    metrics = [
        ("Total Employees", len(users)),
        ("Total Duty Days", sum(r[5] for r in summary_rows)),
        ("Total Leave Days", sum(r[6] for r in summary_rows)),
        ("Vacation Days", sum(r[7] for r in summary_rows)),
        ("Comp-Off Days", sum(r[8] for r in summary_rows)),
        ("Pending Requests", sum(1 for lv in leaves if lv.get("status") == "pending") + sum(1 for s in swaps if str(s.get("status", "")).startswith("pending"))),
        ("Approved Swaps", sum(1 for s in swaps if s.get("status") == "executed")),
    ]
    for idx, (label, value) in enumerate(metrics, start=4):
        dash.cell(idx, 1, label).font = Font(bold=True)
        dash.cell(idx, 2, value)
        dash.cell(idx, 1).fill = PatternFill("solid", fgColor="E5E7EB")
        dash.cell(idx, 2).fill = PatternFill("solid", fgColor="F9FAFB")
    dash["D4"] = "Metric"
    dash["E4"] = "Value"
    for cell in dash[4][3:5]:
        cell.fill = PatternFill("solid", fgColor=colors_x["header"])
        cell.font = Font(color="FFFFFF", bold=True)
    for i, (label, value) in enumerate(metrics[:5], start=5):
        dash.cell(i, 4, label)
        dash.cell(i, 5, value)
    chart = BarChart()
    chart.title = "Workforce Summary"
    chart.y_axis.title = "Count"
    chart.add_data(Reference(dash, min_col=5, min_row=4, max_row=9), titles_from_data=True)
    chart.set_categories(Reference(dash, min_col=4, min_row=5, max_row=9))
    chart.height = 8
    chart.width = 14
    dash.add_chart(chart, "D11")
    dash.column_dimensions["A"].width = 28
    dash.column_dimensions["B"].width = 16

    summary = wb.create_sheet("Monthly Attendance Summary")
    setup_sheet(summary, ["Employee", "Employee ID", "Team", "Role", "Shift", "Duty Days", "Leave Days", "Vacation Days", "Comp-Off Days", "Swap Duty Days", "Absence Days", "Total Hours"])
    for row in summary_rows:
        summary.append(row)
    autofit(summary)

    calendar = wb.create_sheet("Duty Calendar")
    setup_sheet(calendar, ["Employee", "Employee ID", "Team", *date_list])
    for u in users:
        uid = u["_id"]
        row = [u.get("full_name", ""), uid, u.get("team") or ""]
        for d in date_list:
            schedule = sched_by_key.get((uid, d))
            attendance_item = att_by_key.get((uid, d))
            value = ""
            fill = colors_x["duty"]
            leave = next((lv for lv in approved_leaves if lv["user_id"] == uid and lv["start_date"] <= d <= lv["end_date"]), None)
            if leave:
                lt = leave.get("leave_type")
                value = {"annual": "VACATION", "comp_off": "COMP OFF", "sick": "SICK", "emergency": "EMERGENCY"}.get(lt, "LEAVE")
                fill = colors_x["vacation"] if lt == "annual" else colors_x["comp"] if lt == "comp_off" else colors_x["leave"]
            elif schedule:
                value = str(schedule.get("shift_type", "")).upper()
                if attendance_item:
                    value += f" | {str(attendance_item.get('status', '')).upper()}"
            row.append(value)
        calendar.append(row)
        for col in range(4, 4 + len(date_list)):
            value = str(calendar.cell(calendar.max_row, col).value or "")
            fill = colors_x["duty"]
            if "VACATION" in value: fill = colors_x["vacation"]
            elif "COMP OFF" in value: fill = colors_x["comp"]
            elif any(x in value for x in ("SICK", "EMERGENCY", "LEAVE")): fill = colors_x["leave"]
            calendar.cell(calendar.max_row, col).fill = PatternFill("solid", fgColor=fill)
            calendar.cell(calendar.max_row, col).alignment = Alignment(wrap_text=True, horizontal="center")
    calendar.freeze_panes = "D2"
    calendar.column_dimensions["A"].width = 24
    calendar.column_dimensions["B"].width = 38
    calendar.column_dimensions["C"].width = 10
    for col in range(4, 4 + len(date_list)):
        calendar.column_dimensions[get_column_letter(col)].width = 14

    leave_ws = wb.create_sheet("Leave & Vacation Report")
    setup_sheet(leave_ws, ["Employee", "Leave Type", "From Date", "To Date", "Days", "Status", "Reason", "Approved By"])
    for lv in leaves:
        leave_ws.append([lv.get("user_name"), lv.get("leave_type"), lv.get("start_date"), lv.get("end_date"), lv.get("days"), lv.get("status"), lv.get("reason"), lv.get("approved_by")])
        fill = colors_x["pending"] if lv.get("status") == "pending" else colors_x["vacation"] if lv.get("leave_type") == "annual" else colors_x["comp"] if lv.get("leave_type") == "comp_off" else colors_x["leave"]
        for cell in leave_ws[leave_ws.max_row]: cell.fill = PatternFill("solid", fgColor=fill)
    autofit(leave_ws)

    comp_ws = wb.create_sheet("Comp-Off Report")
    setup_sheet(comp_ws, ["Employee", "Earned Date", "Used Date", "Days", "Overtime Hours", "Reason", "Status", "Balance"])
    for item in comp_off:
        is_usage = str(item.get("source_id", "")).startswith("leave-comp-off:")
        user_doc = user_map.get(item.get("user_id"), {})
        comp_ws.append([item.get("user_name"), "" if is_usage else item.get("earned_date"), item.get("earned_date") if is_usage else "", item.get("days"), item.get("overtime_hours", 0), item.get("reason"), "USED" if is_usage else "EARNED", user_doc.get("comp_off_balance", 0)])
    autofit(comp_ws)

    swap_ws = wb.create_sheet("Swap Report")
    setup_sheet(swap_ws, ["Requested By", "Swap Employee", "Date", "Original Shift", "Swap Shift", "Employee Approval", "Admin Approval", "Final Status", "Execution Date", "Reason"])
    for swap in swaps:
        status_value = swap.get("status", "")
        swap_ws.append([
            swap.get("requester_name"), swap.get("swap_user_name"), swap.get("shift_date"), swap.get("requester_original_shift"), swap.get("swap_user_original_shift"),
            "APPROVED" if status_value in ("pending_admin_approval", "executed") else "PENDING" if status_value == "pending_employee_approval" else "REJECTED",
            "APPROVED" if status_value == "executed" else "PENDING" if status_value == "pending_admin_approval" else "-",
            status_value, str(swap.get("executed_at") or ""), swap.get("reason"),
        ])
        fill = colors_x["swap"] if status_value == "executed" else colors_x["pending"] if status_value.startswith("pending") else colors_x["leave"]
        for cell in swap_ws[swap_ws.max_row]: cell.fill = PatternFill("solid", fgColor=fill)
    autofit(swap_ws)

    raw = wb.create_sheet("Raw Attendance Data")
    setup_sheet(raw, ["Date", "Employee", "Employee ID", "Status", "Clock In", "Clock Out", "Hours", "Shift", "Notes", "Marked By"])
    for item in attendance:
        raw.append([item.get("attendance_date"), item.get("user_name"), item.get("user_id"), item.get("status"), item.get("clock_in"), item.get("clock_out"), item.get("hours_worked"), item.get("shift_type"), item.get("notes"), item.get("marked_by")])
    autofit(raw)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"warehouse-management-{start_date}-to-{end_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
